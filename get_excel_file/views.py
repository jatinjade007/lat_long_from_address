"""
views.py - views for accounts app

modification history
--------------------
01a,02jan20,anp  written.

DESCRIPTION
This file contains views for accounts app.
"""

from django.shortcuts import render
from django.http import HttpResponse
import openpyxl
import geopy
from geopy.geocoders import Nominatim
import xlwt

"""
********************************************************************************
*
* download_excel_data - Helps to download the excel data.
* 
* This function will helps to download the excel data.
*
* RETURNS: Excel data.
*
* ERRNO: N/A
*
* SEE_ALSO: N/A.
"""     
def download_excel_data(data):
    # content-type of response
    response = HttpResponse(content_type='application/ms-excel')
    
    #decide file name
    response['Content-Disposition'] = 'attachment; filename="Address.xlsx"'
    
    #creating workbook
    wb = xlwt.Workbook(encoding='utf-8')
    
    #adding sheet
    ws = wb.add_sheet("sheet1")
    
    # Sheet header, first row
    row_num = 0
    
    font_style = xlwt.XFStyle()
    # headers are bold
    font_style.font.bold = True
    
    #column header names, you can use your own headers here
    columns = ['Address', 'Latitude', 'Longitude' ]
    
    #write column headers in sheet
    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style)
      
    font_style = xlwt.XFStyle()  

    for my_row in data:
        row_num = row_num + 1
        ws.write(row_num, 0, my_row['address'], font_style)
        ws.write(row_num, 1, my_row['latitude'], font_style)
        ws.write(row_num, 2, my_row['longitude'], font_style)
    wb.save(response)
    return response

"""
********************************************************************************
*
* index - Helps to request excel file and sends updated excel file in response.
* 
* This function will helps to request excel file and sends updated excel file in response.
*
* RETURNS: Response.
*
* ERRNO: N/A
*
* SEE_ALSO: N/A.
"""     
def index(request):
    if "GET" == request.method:
        return render(request, 'myapp/index.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting all sheets
        sheets = wb.sheetnames
        print(sheets)

        # getting a particular sheet
        worksheet = wb["Sheet1"]
        print(worksheet)

        # getting active sheet
        active_sheet = wb.active
        print(active_sheet)

        # reading a cell
        print(worksheet["A1"].value)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = dict()
            if row[0].value != "Address":
                print(row[0].value)
                locator = Nominatim(user_agent="myGeocoder")
                location = locator.geocode(row[0].value)
                # print("Latitude = {}, Longitude = {}".format(location.latitude, location.longitude))
                row_data["address"]= str(row[0].value)
                row_data["latitude"]= str(location.latitude)
                row_data["longitude"]=str(location.longitude)
                excel_data.append(row_data)
        # print (excel_data)
        response = download_excel_data(excel_data)
        return response









